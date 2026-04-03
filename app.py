import streamlit as st
import pandas as pd
from datetime import datetime, date, timedelta
import uuid
import base64
import time
import json
import io
import urllib.parse
import csv

# ============================================================
# PAGE CONFIG
# ============================================================
st.set_page_config(
    page_title="🛕 Sree Bhadreshwari Amman Temple",
    page_icon="🛕",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================
# SUPABASE CONNECTION
# ============================================================
try:
    from supabase import create_client, Client
    SUPABASE_URL = st.secrets["SUPABASE_URL"]
    SUPABASE_KEY = st.secrets["SUPABASE_KEY"]

    @st.cache_resource
    def get_supabase_client():
        return create_client(SUPABASE_URL, SUPABASE_KEY)

    supabase: Client = get_supabase_client()
    DB_CONNECTED = True
except Exception as e:
    DB_CONNECTED = False
    st.error(f"Database connection failed: {str(e)}")

# ============================================================
# BARCODE
# ============================================================
BARCODE_AVAILABLE = False
try:
    import barcode
    from barcode.writer import ImageWriter
    BARCODE_AVAILABLE = True
except ImportError:
    pass

QRCODE_AVAILABLE = False
try:
    import qrcode
    QRCODE_AVAILABLE = True
except ImportError:
    pass

# ============================================================
# PDF GENERATION
# ============================================================
PDF_AVAILABLE = False
try:
    from fpdf import FPDF

    class BillPDF(FPDF):
        def __init__(self, amman_img_path=None):
            super().__init__()
            self.amman_img_path = amman_img_path

        def header(self):
            if self.amman_img_path:
                try:
                    self.image(self.amman_img_path, 10, 8, 25)
                    self.image(self.amman_img_path, 175, 8, 25)
                except:
                    pass
            self.set_font('Helvetica', 'B', 16)
            self.cell(0, 10, 'Sree Bhadreshwari Amman Temple', 0, 1, 'C')
            self.set_font('Helvetica', '', 10)
            self.cell(0, 6, 'Amme Narayana .. Devi Narayana', 0, 1, 'C')
            self.set_font('Helvetica', '', 8)
            self.cell(0, 5, 'Official Receipt', 0, 1, 'C')
            self.line(10, self.get_y() + 2, 200, self.get_y() + 2)
            self.ln(5)

        def footer(self):
            self.set_y(-30)
            self.line(10, self.get_y(), 200, self.get_y())
            self.ln(3)
            self.set_font('Helvetica', 'I', 8)
            self.cell(0, 5, 'Thank you! May Goddess bless you!', 0, 1, 'C')
            self.cell(0, 5, 'Amme Narayana .. Devi Narayana', 0, 1, 'C')
            self.cell(0, 5, f'Generated: {datetime.now().strftime("%d-%m-%Y %H:%M:%S")}', 0, 1, 'C')

    def save_base64_image_to_temp(base64_str):
        if not base64_str:
            return None
        try:
            import tempfile
            if ',' in base64_str:
                _, data = base64_str.split(',', 1)
            else:
                data = base64_str
            img_data = base64.b64decode(data)
            ext = '.jpg' if ('jpeg' in base64_str or 'jpg' in base64_str) else '.png'
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=ext)
            tmp.write(img_data)
            tmp.close()
            return tmp.name
        except:
            return None

    def generate_bill_pdf(bill_no, manual_bill, bill_book, bill_date,
                          name, address, mobile, pooja_type, amount,
                          amman_base64=None):
        amman_path = None
        if amman_base64 and not amman_base64.startswith('data:image/svg'):
            amman_path = save_base64_image_to_temp(amman_base64)
        pdf = BillPDF(amman_img_path=amman_path)
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=35)
        y_start = pdf.get_y()
        pdf.rect(10, y_start, 190, 100, 'D')
        pdf.set_xy(15, y_start + 5)
        pdf.set_font('Helvetica', 'B', 14)
        pdf.cell(0, 8, 'BILL / RECEIPT', 0, 1, 'C')
        pdf.ln(2)
        for label, value in [("Bill No", str(bill_no or '')),
                              ("Manual Bill No", str(manual_bill or '')),
                              ("Bill Book No", str(bill_book or '')),
                              ("Date", str(bill_date or ''))]:
            pdf.set_x(15)
            pdf.set_font('Helvetica', 'B', 10)
            pdf.cell(50, 7, f"{label}:", 0, 0)
            pdf.set_font('Helvetica', '', 10)
            pdf.cell(0, 7, value, 0, 1)
        pdf.ln(2)
        pdf.line(15, pdf.get_y(), 195, pdf.get_y())
        pdf.ln(3)
        for label, value in [("Name", str(name or '')),
                              ("Address", str(address or '')),
                              ("Mobile", str(mobile or ''))]:
            pdf.set_x(15)
            pdf.set_font('Helvetica', 'B', 10)
            pdf.cell(50, 7, f"{label}:", 0, 0)
            pdf.set_font('Helvetica', '', 10)
            pdf.cell(0, 7, value, 0, 1)
        pdf.ln(5)
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
        pdf.ln(5)
        pdf.set_font('Helvetica', 'B', 12)
        pdf.set_x(15)
        pdf.cell(50, 8, "Pooja Type:", 0, 0)
        pdf.set_font('Helvetica', '', 12)
        pdf.cell(0, 8, str(pooja_type or ''), 0, 1)
        pdf.ln(3)
        pdf.set_font('Helvetica', 'B', 16)
        pdf.set_x(15)
        pdf.cell(50, 12, "Amount:", 0, 0)
        pdf.set_text_color(0, 128, 0)
        pdf.cell(0, 12, f"Rs. {float(amount):,.2f}", 0, 1)
        pdf.set_text_color(0, 0, 0)
        pdf.ln(3)
        pdf.set_font('Helvetica', 'I', 9)
        pdf.set_x(15)
        pdf.cell(0, 6, f"Amount: Rupees {int(float(amount))} Only", 0, 1)
        if amman_path:
            try:
                import os
                os.unlink(amman_path)
            except:
                pass
        return bytes(pdf.output())

    PDF_AVAILABLE = True
except Exception:
    PDF_AVAILABLE = False

# ============================================================
# EXCEL ENGINE
# ============================================================
EXCEL_ENGINE = None
try:
    import xlsxwriter
    EXCEL_ENGINE = 'xlsxwriter'
except ImportError:
    try:
        import openpyxl
        EXCEL_ENGINE = 'openpyxl'
    except ImportError:
        pass

# ============================================================
# CONSTANTS
# ============================================================
NATCHATHIRAM_LIST = [
    "Ashwini", "Bharani", "Karthigai", "Rohini", "Mrigashirsha",
    "Thiruvadirai", "Punarvasu", "Pushya", "Ashlesha", "Magha",
    "Purva Phalguni", "Uttara Phalguni", "Hasta", "Chitra", "Swati",
    "Vishakha", "Anuradha", "Jyeshtha", "Mula", "Purva Ashadha",
    "Uttara Ashadha", "Shravana", "Dhanishta", "Shatabhisha",
    "Purva Bhadrapada", "Uttara Bhadrapada", "Revati"
]
RELATION_TYPES = [
    "Self", "Spouse", "Son", "Daughter", "Father", "Mother",
    "Brother", "Sister", "Grandfather", "Grandmother",
    "Father-in-law", "Mother-in-law", "Son-in-law",
    "Daughter-in-law", "Uncle", "Aunt", "Nephew", "Niece", "Other"
]
MIN_DATE = date(1900, 1, 1)
MAX_DATE = date(2050, 12, 31)

# ============================================================
# DEFAULT AMMAN SVG
# ============================================================
DEFAULT_AMMAN_SVG = """<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 300 300" width="300" height="300">
<defs>
<radialGradient id="glow" cx="50%" cy="50%" r="50%">
<stop offset="0%" style="stop-color:#fff8f0;stop-opacity:1"/>
<stop offset="60%" style="stop-color:#ffe0b2;stop-opacity:1"/>
<stop offset="100%" style="stop-color:#ffcc80;stop-opacity:1"/>
</radialGradient>
</defs>
<circle cx="150" cy="150" r="148" fill="url(#glow)" stroke="#ff6b35" stroke-width="4"/>
<text x="150" y="55" text-anchor="middle" font-size="16" fill="#c62828" font-weight="bold">Om Amman</text>
<text x="150" y="100" text-anchor="middle" font-size="52">🙏</text>
<text x="150" y="140" text-anchor="middle" font-size="40">🪷</text>
<text x="150" y="175" text-anchor="middle" font-size="15" fill="#8B0000" font-weight="bold">Sree Bhadreshwari</text>
<text x="150" y="195" text-anchor="middle" font-size="15" fill="#8B0000" font-weight="bold">Amman</text>
<text x="150" y="220" text-anchor="middle" font-size="10" fill="#c62828">Amme Narayana</text>
<text x="150" y="245" text-anchor="middle" font-size="9" fill="#e65100">Devi Narayana</text>
</svg>"""

AMMAN_IMAGE_BASE64 = "data:image/svg+xml;base64," + base64.b64encode(
    DEFAULT_AMMAN_SVG.strip().encode()).decode()


# ============================================================
# GORGEOUS CSS
# ============================================================
def get_custom_css(amman_bg_url=None):
    bg_overlay = ""
    if amman_bg_url and not amman_bg_url.startswith('data:image/svg'):
        bg_overlay = f"""
        .login-bg-watermark {{
            position: fixed; top: 0; left: 0; width: 100vw; height: 100vh;
            background-image: url('{amman_bg_url}');
            background-size: 300px; background-position: center;
            background-repeat: no-repeat; opacity: 0.06;
            z-index: 0; pointer-events: none;
        }}"""

    return f"""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700;800&display=swap');
        * {{ font-family: 'Poppins', sans-serif; }}
        {bg_overlay}

        .login-container {{
            padding: 40px 35px; border-radius: 24px;
            background: rgba(255, 255, 255, 0.12);
            backdrop-filter: blur(20px);
            -webkit-backdrop-filter: blur(20px);
            box-shadow: 0 20px 60px rgba(0,0,0,0.4), 0 0 40px rgba(255,107,53,0.15),
                        inset 0 1px 0 rgba(255,255,255,0.2);
            border: 1px solid rgba(255,255,255,0.18);
            animation: containerFloat 6s ease-in-out infinite;
        }}
        @keyframes containerFloat {{
            0%, 100% {{ transform: translateY(0px); }}
            50% {{ transform: translateY(-5px); }}
        }}

        .amman-circle {{ text-align: center; margin: 0 auto 25px auto; position: relative; }}
        .amman-circle img {{
            width: 170px; height: 170px; border-radius: 50%; object-fit: cover;
            border: 5px solid transparent;
            background-image: linear-gradient(white, white),
                              linear-gradient(135deg, #ff6b35, #f7c948, #ff6b35, #e91e63);
            background-origin: border-box; background-clip: content-box, border-box;
            box-shadow: 0 0 30px rgba(255,107,53,0.5), 0 0 60px rgba(247,201,72,0.3),
                        0 0 90px rgba(255,107,53,0.15);
            animation: ammanGlow 4s ease-in-out infinite;
        }}
        .amman-outer-ring {{
            position: absolute; top: 50%; left: 50%;
            transform: translate(-50%, -50%);
            width: 200px; height: 200px; border-radius: 50%;
            border: 2px dashed rgba(247,201,72,0.5);
            animation: ringRotate 20s linear infinite;
        }}
        .amman-outer-ring2 {{
            position: absolute; top: 50%; left: 50%;
            transform: translate(-50%, -50%);
            width: 220px; height: 220px; border-radius: 50%;
            border: 1px dotted rgba(255,107,53,0.3);
            animation: ringRotate 30s linear infinite reverse;
        }}
        @keyframes ammanGlow {{
            0% {{ box-shadow: 0 0 30px rgba(255,107,53,0.5), 0 0 60px rgba(247,201,72,0.3); }}
            50% {{ box-shadow: 0 0 40px rgba(233,30,99,0.5), 0 0 80px rgba(255,107,53,0.3), 0 0 120px rgba(247,201,72,0.15); }}
            100% {{ box-shadow: 0 0 30px rgba(255,107,53,0.5), 0 0 60px rgba(247,201,72,0.3); }}
        }}
        @keyframes ringRotate {{
            0% {{ transform: translate(-50%, -50%) rotate(0deg); }}
            100% {{ transform: translate(-50%, -50%) rotate(360deg); }}
        }}

        .temple-name-login {{
            color: #ffffff; font-size: 1.5em; font-weight: 700;
            text-align: center; margin: 10px 0; line-height: 1.3;
            text-shadow: 0 2px 10px rgba(0,0,0,0.3), 0 0 30px rgba(247,201,72,0.3);
        }}
        .tamil-text-login {{
            color: #f7c948; font-size: 1.15em; font-weight: 600;
            text-align: center; margin: 5px 0 25px 0;
            text-shadow: 0 2px 8px rgba(0,0,0,0.4);
            animation: tamilPulse 3s ease-in-out infinite;
        }}
        @keyframes tamilPulse {{ 0%, 100% {{ opacity: 1; }} 50% {{ opacity: 0.7; }} }}

        .login-divider {{
            height: 2px; background: linear-gradient(90deg, transparent, #f7c948, #ff6b35, #f7c948, transparent);
            margin: 20px 0; border: none;
        }}
        .login-footer {{ text-align: center; color: rgba(255,255,255,0.5); font-size: 0.75em; margin-top: 20px; }}

        .deco-lamp {{
            position: fixed; font-size: 2em; animation: floatLamp 8s ease-in-out infinite;
            opacity: 0.3; z-index: 0; pointer-events: none;
        }}
        @keyframes floatLamp {{
            0%, 100% {{ transform: translateY(0) rotate(0deg); opacity: 0.2; }}
            50% {{ transform: translateY(-20px) rotate(10deg); opacity: 0.5; }}
        }}

        .main-header {{
            background: linear-gradient(135deg, #ff6b35 0%, #f7c948 30%, #ffb347 50%, #f7c948 70%, #ff6b35 100%);
            padding: 15px 80px; border-radius: 15px; text-align: center;
            margin-bottom: 20px; box-shadow: 0 4px 20px rgba(255,107,53,0.35);
            position: relative; overflow: hidden;
        }}
        .main-header h1 {{ color: #8B0000; font-size: 1.6em; margin: 0; position: relative; }}
        .main-header p {{ color: #5a1a00; font-size: 0.95em; margin: 5px 0 0 0; position: relative; }}
        .header-amman-left, .header-amman-right {{
            position: absolute; top: 50%; transform: translateY(-50%);
            width: 55px; height: 55px; border-radius: 50%;
            border: 3px solid rgba(139,0,0,0.4); object-fit: cover;
            box-shadow: 0 0 15px rgba(255,107,53,0.3);
        }}
        .header-amman-left {{ left: 15px; }}
        .header-amman-right {{ right: 15px; }}

        div[data-testid="stSidebar"] {{
            background: linear-gradient(180deg, #0f0c29 0%, #1a1a3e 40%, #24243e 70%, #302b63 100%);
        }}
        .sidebar-amman {{ text-align: center; margin: 0 auto 10px auto; }}
        .sidebar-amman img {{
            width: 85px; height: 85px; border-radius: 50%;
            border: 3px solid #f7c948; box-shadow: 0 0 20px rgba(247,201,72,0.4); object-fit: cover;
        }}
        div[data-testid="stSidebar"] .stButton > button {{
            width: 100%; text-align: left;
            background: linear-gradient(135deg, rgba(255,107,53,0.08), rgba(247,201,72,0.08));
            color: #f0f0f0; border: 1px solid rgba(255,255,255,0.08);
            border-radius: 10px; margin: 2px 0; padding: 10px 15px; transition: all 0.3s ease;
        }}
        div[data-testid="stSidebar"] .stButton > button:hover {{
            background: linear-gradient(135deg, rgba(255,107,53,0.35), rgba(247,201,72,0.2));
            border-color: #ff6b35; transform: translateX(3px);
        }}

        .metric-card {{
            padding: 22px; border-radius: 16px; color: white;
            text-align: center; margin: 5px; box-shadow: 0 8px 25px rgba(0,0,0,0.15);
            transition: transform 0.3s ease;
        }}
        .metric-card:hover {{ transform: translateY(-3px); }}
        .metric-card.income {{ background: linear-gradient(135deg, #11998e, #38ef7d); }}
        .metric-card.expense {{ background: linear-gradient(135deg, #eb3349, #f45c43); }}
        .metric-card.balance {{ background: linear-gradient(135deg, #4facfe, #00f2fe); }}
        .metric-card.info {{ background: linear-gradient(135deg, #667eea, #764ba2); }}
        .metric-card h3 {{ margin: 0; font-size: 0.85em; opacity: 0.9; }}
        .metric-card h2 {{ margin: 5px 0 0 0; font-size: 1.7em; }}

        .news-ticker-wrapper {{
            background: linear-gradient(90deg, #1a1a2e, #16213e, #0f3460);
            padding: 12px 20px; border-radius: 12px; overflow: hidden;
            white-space: nowrap; margin: 10px 0; border: 1px solid rgba(247,201,72,0.2);
        }}
        .news-ticker-text {{ display: inline-block; color: #f7c948; font-size: 1em; animation: scroll-left 35s linear infinite; }}
        @keyframes scroll-left {{ 0% {{ transform: translateX(100%); }} 100% {{ transform: translateX(-200%); }} }}

        .pooja-card {{
            background: linear-gradient(135deg, #ffecd2, #fcb69f);
            padding: 12px 15px; border-radius: 12px; margin: 5px 0; border-left: 5px solid #ff6b35;
        }}
        .birthday-card {{
            background: linear-gradient(135deg, #a8edea, #fed6e3);
            padding: 12px 15px; border-radius: 12px; margin: 5px 0; border-left: 5px solid #e91e63;
        }}
        .success-box {{
            background: linear-gradient(135deg, #d4edda, #c3e6cb);
            border: 1px solid #a3d9a5; padding: 15px; border-radius: 12px; color: #155724; margin: 10px 0;
        }}
        .devotee-info-card {{
            background: linear-gradient(135deg, #e8eaf6, #c5cae9);
            padding: 15px; border-radius: 12px; margin: 8px 0;
            border-left: 5px solid #3f51b5;
        }}
        .wa-btn {{
            display: inline-block; background: linear-gradient(135deg, #25D366, #128C7E);
            color: white !important; padding: 12px 28px; border-radius: 12px;
            text-decoration: none; font-weight: 600; font-size: 0.95em; margin: 5px;
            box-shadow: 0 4px 15px rgba(37,211,102,0.35); transition: all 0.3s ease;
        }}
        .wa-btn:hover {{ background: linear-gradient(135deg, #128C7E, #075E54); transform: translateY(-2px); color: white !important; }}
        .wa-btn-small {{
            display: inline-block; background: linear-gradient(135deg, #25D366, #128C7E);
            color: white !important; padding: 6px 14px; border-radius: 8px;
            text-decoration: none; font-weight: 600; font-size: 0.8em; margin: 3px;
        }}
        .wa-btn-small:hover {{ background: linear-gradient(135deg, #128C7E, #075E54); color: white !important; }}

        .upload-error {{ background: #ffebee; border: 1px solid #ef9a9a; padding: 10px; border-radius: 10px; margin: 5px 0; }}
        .barcode-container {{
            background: white; padding: 20px; border-radius: 12px;
            border: 2px dashed #ccc; text-align: center; margin: 10px 0;
        }}
        .barcode-container img {{ max-width: 100%; height: auto; }}

        .bill-receipt {{
            background: linear-gradient(135deg, #fffdf7, #fff8f0);
            padding: 30px; border: 3px solid #ff6b35; border-radius: 20px;
            max-width: 600px; margin: 20px auto;
            box-shadow: 0 8px 30px rgba(255,107,53,0.15);
        }}
        .bill-header {{
            text-align: center; border-bottom: 3px solid #ff6b35;
            padding-bottom: 15px; position: relative;
        }}

        .scan-preview {{
            background: #f5f5f5; border: 2px solid #ddd; border-radius: 12px;
            padding: 10px; text-align: center; margin: 8px 0;
        }}
        .scan-preview img {{
            max-width: 100%; max-height: 300px; border-radius: 8px;
        }}
    </style>"""


# ============================================================
# SESSION STATE
# ============================================================
defaults = {
    'logged_in': False, 'username': '', 'user_role': '',
    'current_page': 'Dashboard', 'custom_amman_photo': None,
}
for key, val in defaults.items():
    if key not in st.session_state:
        st.session_state[key] = val


# ============================================================
# AMMAN IMAGE MANAGEMENT
# ============================================================
def get_amman_image():
    if DB_CONNECTED:
        try:
            settings = db_select("temple_settings", filters={"key": "amman_image"})
            if settings and settings[0].get('value'):
                return settings[0]['value']
        except:
            pass
    if st.session_state.get('custom_amman_photo'):
        return st.session_state['custom_amman_photo']
    return AMMAN_IMAGE_BASE64

def save_amman_image_to_db(base64_img):
    if not DB_CONNECTED:
        st.session_state['custom_amman_photo'] = base64_img
        return False
    try:
        existing = db_select("temple_settings", filters={"key": "amman_image"})
        if existing:
            db_update("temple_settings", {"value": base64_img}, "key", "amman_image")
        else:
            db_insert("temple_settings", {"key": "amman_image", "value": base64_img})
        st.session_state['custom_amman_photo'] = base64_img
        return True
    except:
        st.session_state['custom_amman_photo'] = base64_img
        return False

def get_amman_for_pdf():
    img = get_amman_image()
    return img if img and not img.startswith('data:image/svg') else None


# ============================================================
# DATABASE HELPERS
# ============================================================
def db_select(table, columns="*", filters=None, gte_filters=None, lte_filters=None):
    try:
        query = supabase.table(table).select(columns)
        if filters:
            for k, v in filters.items():
                query = query.eq(k, v)
        if gte_filters:
            for k, v in gte_filters.items():
                query = query.gte(k, str(v))
        if lte_filters:
            for k, v in lte_filters.items():
                query = query.lte(k, str(v))
        result = query.execute()
        return result.data if result.data else []
    except:
        return []

def db_insert(table, data):
    try:
        result = supabase.table(table).insert(data).execute()
        return result.data if result.data else None
    except Exception as e:
        st.error(f"Insert Error ({table}): {e}")
        return None

def db_update(table, data, col, val):
    try:
        return supabase.table(table).update(data).eq(col, val).execute().data
    except:
        return None

def db_delete(table, col, val):
    try:
        supabase.table(table).delete().eq(col, val).execute()
        return True
    except:
        return False

def file_to_base64(f):
    if f:
        return f"data:{f.type};base64,{base64.b64encode(f.getvalue()).decode()}"
    return None

def get_income(s, e):
    return sum(float(b.get('amount', 0)) for b in db_select("bills", "amount", gte_filters={"bill_date": s}, lte_filters={"bill_date": e}))

def get_expense(s, e):
    return sum(float(x.get('amount', 0)) for x in db_select("expenses", "amount", gte_filters={"expense_date": s}, lte_filters={"expense_date": e}))

def get_period_dates(p):
    t = date.today()
    if p == "Daily": return t, t
    elif p == "Weekly": return t - timedelta(days=t.weekday()), t
    elif p == "Monthly": return t.replace(day=1), t
    elif p == "Yearly": return t.replace(month=1, day=1), t
    return t, t

def get_todays_birthdays():
    t = date.today()
    bdays = []
    for d in db_select("devotees", "name, dob"):
        if d.get('dob'):
            try:
                dob = datetime.strptime(str(d['dob']), '%Y-%m-%d').date()
                if dob.month == t.month and dob.day == t.day:
                    bdays.append(f"🎂 {d['name']} (Devotee)")
            except:
                pass
    for m in db_select("family_members", "name, dob"):
        if m.get('dob'):
            try:
                dob = datetime.strptime(str(m['dob']), '%Y-%m-%d').date()
                if dob.month == t.month and dob.day == t.day:
                    bdays.append(f"🎂 {m['name']} (Family)")
            except:
                pass
    return bdays

def gen_bill_no():
    return f"TMS-{datetime.now().strftime('%Y%m%d%H%M%S')}-{str(uuid.uuid4())[:4].upper()}"

def make_whatsapp_link(phone, message):
    phone_clean = ''.join(filter(str.isdigit, str(phone)))
    if len(phone_clean) == 10:
        phone_clean = "91" + phone_clean
    return f"https://wa.me/{phone_clean}?text={urllib.parse.quote(message)}"

def parse_date_safe(val):
    if val is None or str(val).strip() == '' or str(val).lower() in ('nan', 'nat', 'none'):
        return None
    val_str = str(val).strip()
    for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d']:
        try:
            return datetime.strptime(val_str, fmt).date()
        except:
            pass
    return None

def safe_str(val):
    if val is None: return ''
    s = str(val).strip()
    return '' if s.lower() in ('nan', 'none', 'nat') else s


# ============================================================
# BARCODE HELPERS
# ============================================================
def generate_barcode_image(data_str, barcode_type='code128'):
    if BARCODE_AVAILABLE:
        try:
            barcode_class = barcode.get_barcode_class(barcode_type)
            buffer = io.BytesIO()
            b = barcode_class(str(data_str), writer=ImageWriter())
            b.write(buffer, options={'module_width': 0.4, 'module_height': 15, 'font_size': 10})
            buffer.seek(0)
            img_base64 = base64.b64encode(buffer.getvalue()).decode()
            return f"data:image/png;base64,{img_base64}", buffer.getvalue()
        except:
            pass
    if QRCODE_AVAILABLE:
        try:
            qr = qrcode.QRCode(version=1, box_size=10, border=5)
            qr.add_data(str(data_str))
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            buffer = io.BytesIO()
            img.save(buffer, format='PNG')
            buffer.seek(0)
            img_base64 = base64.b64encode(buffer.getvalue()).decode()
            return f"data:image/png;base64,{img_base64}", buffer.getvalue()
        except:
            pass
    import hashlib
    h = hashlib.md5(str(data_str).encode()).hexdigest()
    bars, x = [], 10
    for c in h[:32]:
        v = int(c, 16)
        w = 2 if v > 7 else 1
        if v % 2 == 0:
            bars.append(f'<rect x="{x}" y="10" width="{w}" height="60" fill="black"/>')
        x += w + 1
    tw = x + 10
    svg = f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {tw} 90" width="{tw}" height="90"><rect width="{tw}" height="90" fill="white"/>{"".join(bars)}<text x="{tw/2}" y="82" text-anchor="middle" font-size="8" font-family="monospace">{data_str}</text></svg>'
    return "data:image/svg+xml;base64," + base64.b64encode(svg.encode()).decode(), None

def generate_asset_barcode_pdf(asset_tag, asset_name, barcode_img_bytes=None):
    if not PDF_AVAILABLE:
        return None
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 14)
    pdf.cell(0, 10, 'Sree Bhadreshwari Amman Temple', 0, 1, 'C')
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 6, 'Asset Barcode Label', 0, 1, 'C')
    pdf.ln(10)
    pdf.set_font('Helvetica', 'B', 12)
    pdf.cell(0, 8, f'Asset Tag: {asset_tag}', 0, 1, 'C')
    pdf.cell(0, 8, f'Asset Name: {asset_name}', 0, 1, 'C')
    pdf.ln(5)
    if barcode_img_bytes:
        try:
            import tempfile, os
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
            tmp.write(barcode_img_bytes)
            tmp.close()
            pdf.image(tmp.name, x=40, w=130)
            os.unlink(tmp.name)
        except:
            pdf.set_font('Courier', 'B', 16)
            pdf.cell(0, 10, f'|| {asset_tag} ||', 0, 1, 'C')
    pdf.ln(5)
    pdf.set_font('Helvetica', 'I', 8)
    pdf.cell(0, 5, f'Generated: {datetime.now().strftime("%d-%m-%Y %H:%M")}', 0, 1, 'C')
    return bytes(pdf.output())


# ============================================================
# COMMON HEADER & BULK UPLOAD
# ============================================================
def render_page_header(title, subtitle=""):
    amman_img = get_amman_image()
    st.markdown(f"""
    <div class="main-header">
        <img src="{amman_img}" class="header-amman-left" alt="">
        <h1>{title}</h1>
        <p>{subtitle}</p>
        <img src="{amman_img}" class="header-amman-right" alt="">
    </div>""", unsafe_allow_html=True)

def generate_bulk_template():
    columns = ['Sl_No', 'Type', 'Family_Head_Name', 'Member_Name', 'Address',
               'Mobile_No', 'WhatsApp_No', 'Relation_Type', 'Date_of_Birth',
               'Natchathiram', 'Wedding_Day', 'Yearly_Pooja', 'Yearly_Pooja_Dates']
    sample = [
        ['1', 'HEAD', 'Raman K', '', '12 Main St', '9876543210', '9876543210', 'Self', '15-05-1980', 'Ashwini', '10-06-2005', 'Archana;Abhishekam', '15-01-2025;20-06-2025'],
        ['1.1', 'MEMBER', 'Raman K', 'Lakshmi R', '', '', '', 'Spouse', '20-07-1985', 'Bharani', '10-06-2005', '', ''],
    ]
    df = pd.DataFrame(sample, columns=columns)
    if EXCEL_ENGINE:
        try:
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine=EXCEL_ENGINE) as writer:
                df.to_excel(writer, index=False, sheet_name='Devotees')
            return output.getvalue(), 'template.xlsx', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        except:
            pass
    output = io.StringIO()
    df.to_csv(output, index=False)
    return output.getvalue().encode('utf-8'), 'template.csv', 'text/csv'

def process_bulk_upload(df):
    results = {'success': 0, 'errors': [], 'members_added': 0, 'poojas_added': 0}
    head_id_map = {}
    df.columns = [c.strip().replace(' ', '_') for c in df.columns]
    for col in ['Type', 'Family_Head_Name']:
        if col not in df.columns:
            results['errors'].append(f"Missing: {col}")
            return results
    for idx, row in df[df['Type'].astype(str).str.upper().str.strip() == 'HEAD'].iterrows():
        try:
            name = safe_str(row.get('Family_Head_Name'))
            if not name:
                continue
            dob = parse_date_safe(row.get('Date_of_Birth'))
            wed = parse_date_safe(row.get('Wedding_Day'))
            r = db_insert("devotees", {"name": name, "dob": str(dob) if dob else None,
                "relation_type": safe_str(row.get('Relation_Type')) or 'Self',
                "mobile_no": safe_str(row.get('Mobile_No')), "whatsapp_no": safe_str(row.get('WhatsApp_No')),
                "wedding_day": str(wed) if wed else None, "natchathiram": safe_str(row.get('Natchathiram')) or None,
                "address": safe_str(row.get('Address'))})
            if r:
                hid = r[0]['id']
                head_id_map[name.lower().strip()] = hid
                results['success'] += 1
                ps = safe_str(row.get('Yearly_Pooja'))
                if ps:
                    ds = safe_str(row.get('Yearly_Pooja_Dates'))
                    for i, pn in enumerate([p.strip() for p in ps.split(';') if p.strip()]):
                        pd_list = [d.strip() for d in ds.split(';') if d.strip()] if ds else []
                        pd_val = parse_date_safe(pd_list[i]) if i < len(pd_list) else None
                        db_insert("devotee_yearly_pooja", {"devotee_id": hid, "pooja_type": pn, "pooja_date": str(pd_val) if pd_val else None, "description": "Bulk"})
                        results['poojas_added'] += 1
        except Exception as e:
            results['errors'].append(f"Row {idx+2}: {e}")
    for idx, row in df[df['Type'].astype(str).str.upper().str.strip() == 'MEMBER'].iterrows():
        try:
            href = safe_str(row.get('Family_Head_Name')).lower().strip()
            mname = safe_str(row.get('Member_Name')) or f"Member of {href}"
            hid = head_id_map.get(href)
            if not hid:
                for d in db_select("devotees", "id, name"):
                    if d['name'].lower().strip() == href:
                        hid = d['id']
                        break
            if not hid:
                results['errors'].append(f"Row {idx+2}: Head not found")
                continue
            dob = parse_date_safe(row.get('Date_of_Birth'))
            wed = parse_date_safe(row.get('Wedding_Day'))
            if db_insert("family_members", {"devotee_id": hid, "name": mname, "dob": str(dob) if dob else None,
                "relation_type": safe_str(row.get('Relation_Type')), "wedding_day": str(wed) if wed else None,
                "natchathiram": safe_str(row.get('Natchathiram')) or None}):
                results['members_added'] += 1
        except Exception as e:
            results['errors'].append(f"Row {idx+2}: {e}")
    return results


# ============================================================
# DEVOTEE SEARCH WIDGET (Reusable dropdown with multi-field search)
# ============================================================
def devotee_search_widget(key_prefix="billing"):
    """
    Enhanced devotee search with dropdown.
    Returns: (devotee_id, name, address, mobile, whatsapp) or Nones
    """
    st.markdown("#### 🔍 Search Devotee")

    search_by = st.selectbox(
        "Search By",
        ["Name", "Mobile No", "WhatsApp No", "Address"],
        key=f"{key_prefix}_search_by"
    )

    field_map = {
        "Name": "name",
        "Mobile No": "mobile_no",
        "WhatsApp No": "whatsapp_no",
        "Address": "address"
    }
    db_field = field_map[search_by]

    search_val = st.text_input(
        f"🔎 Enter {search_by}",
        key=f"{key_prefix}_search_val",
        placeholder=f"Type {search_by.lower()} to search..."
    )

    all_devotees = db_select("devotees")

    if search_val and search_val.strip():
        filtered = [
            d for d in all_devotees
            if search_val.lower() in str(d.get(db_field, '')).lower()
        ]
    else:
        filtered = all_devotees

    if not filtered:
        st.warning("⚠️ No devotees found. Try different search.")
        return None, None, None, None, None

    # Build dropdown options showing multiple fields
    dropdown_options = {}
    for d in filtered:
        label = (
            f"👤 {d.get('name', 'N/A')} | "
            f"📱 {d.get('mobile_no', 'N/A')} | "
            f"📲 {d.get('whatsapp_no', 'N/A')} | "
            f"🏠 {(d.get('address', '') or '')[:30]}"
        )
        dropdown_options[label] = d

    selected_label = st.selectbox(
        "📋 Select Devotee from List",
        list(dropdown_options.keys()),
        key=f"{key_prefix}_dropdown"
    )

    if selected_label:
        dev = dropdown_options[selected_label]

        # Show selected devotee info card
        st.markdown(f"""
        <div class="devotee-info-card">
            <table style="width:100%;">
                <tr>
                    <td style="padding:4px;"><b>👤 Name:</b></td>
                    <td style="padding:4px;">{dev.get('name', 'N/A')}</td>
                </tr>
                <tr>
                    <td style="padding:4px;"><b>📱 Mobile:</b></td>
                    <td style="padding:4px;">{dev.get('mobile_no', 'N/A')}</td>
                </tr>
                <tr>
                    <td style="padding:4px;"><b>📲 WhatsApp:</b></td>
                    <td style="padding:4px;">{dev.get('whatsapp_no', 'N/A')}</td>
                </tr>
                <tr>
                    <td style="padding:4px;"><b>🏠 Address:</b></td>
                    <td style="padding:4px;">{dev.get('address', 'N/A')}</td>
                </tr>
                <tr>
                    <td style="padding:4px;"><b>⭐ Star:</b></td>
                    <td style="padding:4px;">{dev.get('natchathiram', 'N/A')}</td>
                </tr>
            </table>
        </div>
        """, unsafe_allow_html=True)

        return (
            dev['id'],
            dev.get('name', ''),
            dev.get('address', ''),
            dev.get('mobile_no', ''),
            dev.get('whatsapp_no', '')
        )

    return None, None, None, None, None


# ============================================================
# WHATSAPP BILL MESSAGE BUILDER
# ============================================================
def build_bill_whatsapp_message(bill_no, bill_date, name, pooja, amount, manual_bill="", book_no=""):
    return (
        f"🛕 *Sree Bhadreshwari Amman Temple*\n"
        f"🙏 அம்மே நாராயணா ..தேவி நாராயணா\n"
        f"\n"
        f"📋 *BILL / RECEIPT*\n"
        f"━━━━━━━━━━━━━━━━━\n"
        f"📄 Bill No: {bill_no}\n"
        f"{'📝 Manual Bill: ' + str(manual_bill) + chr(10) if manual_bill else ''}"
        f"{'📖 Book No: ' + str(book_no) + chr(10) if book_no else ''}"
        f"📅 Date: {bill_date}\n"
        f"━━━━━━━━━━━━━━━━━\n"
        f"👤 Name: {name}\n"
        f"🙏 Pooja: {pooja}\n"
        f"💰 *Amount: ₹ {float(amount):,.2f}*\n"
        f"━━━━━━━━━━━━━━━━━\n"
        f"\n"
        f"🙏 Thank you for your contribution!\n"
        f"May Goddess Bhadreshwari bless you!\n"
        f"\n"
        f"🪔 அம்மே நாராயணா ..தேவி நாராயணா 🪔"
    )


# ============================================================
# PAGE: LOGIN
# ============================================================
def page_login():
    amman_img = get_amman_image()
    st.markdown(get_custom_css(amman_img), unsafe_allow_html=True)

    st.markdown("""
    <style>
        .stApp {
            background: linear-gradient(135deg,
                #0f0c29 0%, #1a1a3e 15%, #302b63 30%, #4a1942 45%,
                #6b2fa0 55%, #8b3a62 65%, #b7472a 78%, #e65100 88%,
                #ff6b35 95%, #f7c948 100%) !important;
        }
    </style>
    """, unsafe_allow_html=True)

    # Floating lamps
    st.markdown("""
    <div class="deco-lamp" style="top:10%;left:5%;">🪔</div>
    <div class="deco-lamp" style="top:20%;right:8%;animation-delay:2s;">🕉️</div>
    <div class="deco-lamp" style="top:60%;left:3%;animation-delay:4s;">🪷</div>
    <div class="deco-lamp" style="top:70%;right:5%;animation-delay:1s;">🪔</div>
    <div class="deco-lamp" style="top:40%;left:8%;animation-delay:3s;">✨</div>
    <div class="deco-lamp" style="top:85%;right:10%;animation-delay:5s;">🕉️</div>
    """, unsafe_allow_html=True)

    if amman_img and not amman_img.startswith('data:image/svg'):
        st.markdown('<div class="login-bg-watermark"></div>', unsafe_allow_html=True)

    c1, c2, c3 = st.columns([1, 2, 1])
    with c2:
        st.markdown('<div class="login-container">', unsafe_allow_html=True)
        st.markdown(f"""
        <div class="amman-circle">
            <div class="amman-outer-ring2"></div>
            <div class="amman-outer-ring"></div>
            <img src="{amman_img}" alt="Sree Bhadreshwari Amman">
        </div>
        """, unsafe_allow_html=True)

        st.markdown("""
        <div class="temple-name-login">
            🛕 Sree Bhadreshwari Amman Temple<br>
            <span style="font-size:0.65em;font-weight:400;opacity:0.8;">Management System</span>
        </div>
        <div class="tamil-text-login">🙏 அம்மே நாராயணா ..தேவி நாராயணா 🙏</div>
        """, unsafe_allow_html=True)

        st.markdown('<div class="login-divider"></div>', unsafe_allow_html=True)

        with st.form("login"):
            u = st.text_input("👤 Username", placeholder="Enter username")
            p = st.text_input("🔑 Password", type="password", placeholder="Enter password")
            st.markdown("<br>", unsafe_allow_html=True)
            if st.form_submit_button("🚀 Enter Temple Portal", use_container_width=True):
                if not u or not p:
                    st.warning("⚠️ Enter both fields!")
                elif not DB_CONNECTED:
                    st.error("❌ Database not connected!")
                else:
                    users = db_select("users", filters={"username": u})
                    if users and users[0].get('password_hash') == p:
                        st.session_state.logged_in = True
                        st.session_state.username = u
                        st.session_state.user_role = users[0].get('role', 'user')
                        st.success("✅ Welcome! Amme Narayana!")
                        time.sleep(0.8)
                        st.rerun()
                    else:
                        st.error("❌ Invalid credentials!")

        st.markdown('<div class="login-divider"></div>', unsafe_allow_html=True)
        st.markdown("""
        <div class="login-footer">🪔 Default: admin / admin123 🪔<br>
        <span style="font-size:0.9em;opacity:0.7;">v3.1 • Temple Management System</span></div>
        """, unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)


# ============================================================
# PAGE: DASHBOARD
# ============================================================
def page_dashboard():
    render_page_header("🛕 Sree Bhadreshwari Amman Temple", "🙏 அம்மே நாராயணா ..தேவி நாராயணா 🙏")

    tparts = get_todays_birthdays()
    for n in db_select("news_ticker", filters={"is_active": True}):
        tparts.append(f"📢 {n['message']}")
    if not tparts:
        tparts.append("🛕 Welcome to Sree Bhadreshwari Amman Temple! 🙏")
    st.markdown(f'<div class="news-ticker-wrapper"><div class="news-ticker-text">{" &nbsp;⭐&nbsp; ".join(tparts)}</div></div>', unsafe_allow_html=True)

    period = st.selectbox("📅 Period", ["Daily", "Weekly", "Monthly", "Yearly"])
    s, e = get_period_dates(period)
    inc, exp = get_income(s, e), get_expense(s, e)
    bal, td = inc - exp, len(db_select("devotees", "id"))

    c1, c2, c3, c4 = st.columns(4)
    with c1: st.markdown(f'<div class="metric-card income"><h3>💰 {period} Income</h3><h2>₹{inc:,.2f}</h2></div>', unsafe_allow_html=True)
    with c2: st.markdown(f'<div class="metric-card expense"><h3>💸 {period} Expenses</h3><h2>₹{exp:,.2f}</h2></div>', unsafe_allow_html=True)
    with c3: st.markdown(f'<div class="metric-card balance"><h3>💎 Balance</h3><h2>₹{bal:,.2f}</h2></div>', unsafe_allow_html=True)
    with c4: st.markdown(f'<div class="metric-card info"><h3>👥 Devotees</h3><h2>{td}</h2></div>', unsafe_allow_html=True)

    st.markdown("---")
    cl, cr = st.columns(2)
    with cl:
        st.markdown("### 🎂 Birthdays Today")
        bdays = get_todays_birthdays()
        for b in bdays:
            st.markdown(f'<div class="birthday-card">🎉 {b} 🎈</div>', unsafe_allow_html=True)
        if not bdays:
            st.info("No birthdays today")
    with cr:
        st.markdown("### 🙏 Today's Pooja")
        for p in db_select("daily_pooja", filters={"pooja_date": str(date.today())}):
            ic = "✅" if p.get('status') == 'completed' else "⏳"
            st.markdown(f'<div class="pooja-card">{ic} <b>{p["pooja_name"]}</b> — {p.get("pooja_time","")}</div>', unsafe_allow_html=True)
            if p.get('status') != 'completed':
                if st.button("Complete", key=f"c_{p['id']}"):
                    db_update("daily_pooja", {"status": "completed"}, "id", p['id'])
                    st.rerun()
        with st.expander("➕ Add Pooja"):
            with st.form("adp"):
                dn = st.text_input("Name")
                dt_t = st.text_input("Time")
                dd = st.date_input("Date")
                if st.form_submit_button("Add"):
                    if dn:
                        db_insert("daily_pooja", {"pooja_name": dn, "pooja_time": dt_t, "pooja_date": str(dd), "status": "pending"})
                        st.rerun()

    st.markdown("---")
    st.bar_chart(pd.DataFrame({"Category": ["Income", "Expenses", "Balance"], "₹": [inc, exp, bal]}).set_index("Category"))


# ============================================================
# PAGE: DEVOTEE ENROLLMENT
# ============================================================
def page_devotee_enrollment():
    render_page_header("👥 Devotee Enrollment", "Register, Bulk Upload & Manage")
    tab1, tab2, tab3, tab4 = st.tabs(["➕ New", "📤 Bulk Upload", "🔍 Search", "👨‍👩‍👧‍👦 Family"])

    with tab1:
        with st.form("enroll", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                nm = st.text_input("👤 Name *")
                db_v = st.date_input("📅 DOB", value=date(1990, 1, 1), min_value=MIN_DATE, max_value=MAX_DATE)
                rl = st.selectbox("👪 Relation", RELATION_TYPES)
                mb = st.text_input("📱 Mobile")
                wa = st.text_input("📲 WhatsApp")
            with c2:
                wd = st.date_input("💒 Wedding", value=None, min_value=MIN_DATE, max_value=MAX_DATE)
                nt = st.selectbox("⭐ Star", ["--"] + NATCHATHIRAM_LIST)
                ad = st.text_area("🏠 Address", height=80)
                ph = st.file_uploader("📷 Photo", type=['jpg', 'jpeg', 'png'])
            st.markdown("#### 🙏 Yearly Pooja")
            yc1, yc2, yc3 = st.columns(3)
            ptl = [p['name'] for p in db_select("pooja_types", "name")]
            with yc1: ypt = st.selectbox("Type", ["--"] + ptl, key="y1t")
            with yc2: ypd = st.date_input("Date", key="y1d", min_value=MIN_DATE, max_value=MAX_DATE)
            with yc3: ypdesc = st.text_input("Desc", key="y1dc")
            if st.form_submit_button("✅ Register", use_container_width=True):
                if nm.strip():
                    r = db_insert("devotees", {
                        "name": nm.strip(), "dob": str(db_v), "relation_type": rl,
                        "mobile_no": mb, "whatsapp_no": wa,
                        "wedding_day": str(wd) if wd else None,
                        "natchathiram": nt if nt != "--" else None,
                        "address": ad, "photo_url": file_to_base64(ph)
                    })
                    if r and ypt != "--":
                        db_insert("devotee_yearly_pooja", {"devotee_id": r[0]['id'], "pooja_type": ypt, "pooja_date": str(ypd), "description": ypdesc})
                    if r:
                        st.success(f"✅ '{nm}' enrolled!")
                        st.rerun()

    with tab2:
        tb, tn, tm = generate_bulk_template()
        st.download_button("📥 Download Template", data=tb, file_name=tn, mime=tm, use_container_width=True)
        uf = st.file_uploader("📁 Upload", type=['xlsx', 'xls', 'csv'], key="bulk")
        if uf:
            try:
                df = pd.read_csv(uf) if uf.name.endswith('.csv') else pd.read_excel(uf, sheet_name=0)
                st.dataframe(df.head(15), use_container_width=True, hide_index=True)
                if st.button("🚀 Process", use_container_width=True, type="primary"):
                    with st.spinner("Processing..."):
                        res = process_bulk_upload(df)
                    rc1, rc2, rc3 = st.columns(3)
                    with rc1: st.markdown(f'<div class="metric-card income"><h3>Heads</h3><h2>{res["success"]}</h2></div>', unsafe_allow_html=True)
                    with rc2: st.markdown(f'<div class="metric-card balance"><h3>Members</h3><h2>{res["members_added"]}</h2></div>', unsafe_allow_html=True)
                    with rc3: st.markdown(f'<div class="metric-card info"><h3>Poojas</h3><h2>{res["poojas_added"]}</h2></div>', unsafe_allow_html=True)
                    if res['errors']:
                        with st.expander(f"⚠️ {len(res['errors'])} Errors"):
                            for err in res['errors']:
                                st.markdown(f'<div class="upload-error">❌ {err}</div>', unsafe_allow_html=True)
                    if res['success'] > 0:
                        st.balloons()
            except Exception as e:
                st.error(f"Error: {e}")

    with tab3:
        sc1, sc2, sc3 = st.columns(3)
        with sc1: sn = st.text_input("Name", key="sn")
        with sc2: sm = st.text_input("Mobile", key="sm")
        with sc3: sa = st.text_input("Address", key="sa")
        devs = db_select("devotees")
        if sn: devs = [d for d in devs if sn.lower() in d.get('name', '').lower()]
        if sm: devs = [d for d in devs if sm in d.get('mobile_no', '')]
        if sa: devs = [d for d in devs if sa.lower() in d.get('address', '').lower()]
        st.markdown(f"**Found: {len(devs)}**")
        for dev in devs:
            with st.expander(f"👤 {dev['name']} | 📱 {dev.get('mobile_no', 'N/A')}"):
                dc1, dc2 = st.columns([3, 1])
                with dc1:
                    for l, k in [("Name", "name"), ("DOB", "dob"), ("Mobile", "mobile_no"), ("WhatsApp", "whatsapp_no"), ("Star", "natchathiram"), ("Address", "address")]:
                        st.write(f"**{l}:** {dev.get(k, 'N/A')}")
                with dc2:
                    if dev.get('photo_url') and dev['photo_url'].startswith('data:'):
                        st.markdown(f'<img src="{dev["photo_url"]}" width="120" style="border-radius:10px">', unsafe_allow_html=True)
                for yp in db_select("devotee_yearly_pooja", filters={"devotee_id": dev['id']}):
                    yc1, yc2 = st.columns([5, 1])
                    with yc1: st.write(f"• {yp['pooja_type']} — {yp.get('pooja_date', '')}")
                    with yc2:
                        if st.button("❌", key=f"dyp_{yp['id']}"):
                            db_delete("devotee_yearly_pooja", "id", yp['id'])
                            st.rerun()
                bc1, bc2 = st.columns(2)
                with bc1:
                    if st.button("✏️ Edit", key=f"e_{dev['id']}"):
                        st.session_state[f"ed_{dev['id']}"] = not st.session_state.get(f"ed_{dev['id']}", False)
                        st.rerun()
                with bc2:
                    if st.button("🗑️ Delete", key=f"d_{dev['id']}"):
                        db_delete("devotee_yearly_pooja", "devotee_id", dev['id'])
                        db_delete("family_members", "devotee_id", dev['id'])
                        db_delete("devotees", "id", dev['id'])
                        st.rerun()
                if st.session_state.get(f"ed_{dev['id']}", False):
                    with st.form(f"ef_{dev['id']}"):
                        en = st.text_input("Name", value=dev.get('name', ''), key=f"en_{dev['id']}")
                        em = st.text_input("Mobile", value=dev.get('mobile_no', ''), key=f"em_{dev['id']}")
                        ewa = st.text_input("WhatsApp", value=dev.get('whatsapp_no', ''), key=f"ewa_{dev['id']}")
                        ea = st.text_area("Address", value=dev.get('address', ''), key=f"ea_{dev['id']}")
                        if st.form_submit_button("💾 Save"):
                            db_update("devotees", {"name": en, "mobile_no": em, "whatsapp_no": ewa, "address": ea}, "id", dev['id'])
                            st.session_state[f"ed_{dev['id']}"] = False
                            st.rerun()

    with tab4:
        ds = db_select("devotees", "id,name,mobile_no")
        if not ds:
            st.info("No devotees")
            return
        do = {f"{d['name']} ({d.get('mobile_no', '')})": d['id'] for d in ds}
        sh = st.selectbox("Head", list(do.keys()))
        hi = do[sh]
        for fm in db_select("family_members", filters={"devotee_id": hi}):
            fc1, fc2 = st.columns([5, 1])
            with fc1: st.write(f"👤 **{fm['name']}** | {fm.get('relation_type', '')} | {fm.get('dob', '')}")
            with fc2:
                if st.button("🗑️", key=f"dfm_{fm['id']}"):
                    db_delete("family_members", "id", fm['id'])
                    st.rerun()
        with st.form("afm", clear_on_submit=True):
            fc1, fc2 = st.columns(2)
            with fc1:
                fn = st.text_input("Name *")
                fd = st.date_input("DOB", value=date(1995, 1, 1), min_value=MIN_DATE, max_value=MAX_DATE)
                fr = st.selectbox("Relation", RELATION_TYPES)
            with fc2:
                fw = st.date_input("Wedding", value=None, min_value=MIN_DATE, max_value=MAX_DATE, key="fmw")
                fs = st.selectbox("Star", ["--"] + NATCHATHIRAM_LIST, key="fms")
            if st.form_submit_button("➕ Add Member", use_container_width=True):
                if fn.strip():
                    db_insert("family_members", {
                        "devotee_id": hi, "name": fn.strip(), "dob": str(fd),
                        "relation_type": fr, "wedding_day": str(fw) if fw else None,
                        "natchathiram": fs if fs != "--" else None
                    })
                    st.rerun()


# ============================================================
# PAGE: BILLING (Enhanced with Devotee Search Dropdown + WhatsApp)
# ============================================================
def page_billing():
    render_page_header("🧾 Billing", "PDF Download & WhatsApp Send")

    tab1, tab2 = st.tabs(["➕ New Bill", "📋 Bill History"])

    with tab1:
        dt = st.radio("Devotee Type", ["Enrolled Devotee", "Guest"], horizontal=True)

        st.markdown("---")

        bc1, bc2 = st.columns(2)

        with bc1:
            st.markdown("#### 📋 Bill Details")
            mbl = st.text_input("📝 Manual Bill No", key="bill_manual")
            bb = st.text_input("📖 Bill Book No", key="bill_book")
            ptd = db_select("pooja_types")
            pto = {f"{p['name']} — ₹{p.get('amount', 0)}": p for p in ptd} if ptd else {}
            sp = st.selectbox("🙏 Pooja Type", list(pto.keys()) if pto else ["None"], key="bill_pooja")
            da = float(pto[sp].get('amount', 0)) if sp in pto else 0.0
            am = st.number_input("💰 Amount (₹)", value=da, min_value=0.0, step=10.0, key="bill_amt")
            bd = st.date_input("📅 Bill Date", value=date.today(), key="bill_date")

        with bc2:
            did = None
            d_name = d_addr = d_mob = d_wa = ""

            if dt == "Enrolled Devotee":
                # ENHANCED DEVOTEE SEARCH WITH DROPDOWN
                did, d_name, d_addr, d_mob, d_wa = devotee_search_widget("billing")

            else:
                st.markdown("#### 👤 Guest Details")
                d_name = st.text_input("👤 Guest Name *", key="guest_name")
                d_addr = st.text_area("🏠 Guest Address", height=60, key="guest_addr")
                d_mob = st.text_input("📱 Mobile No", key="guest_mob")
                d_wa = st.text_input("📲 WhatsApp No", key="guest_wa")

        st.markdown("---")

        # GENERATE BILL BUTTON
        if st.button("🧾 Generate Bill", use_container_width=True, type="primary"):
            ok = True
            if dt == "Enrolled Devotee" and not did:
                st.error("❌ Please search and select a devotee!")
                ok = False
            if dt == "Guest" and not d_name.strip():
                st.error("❌ Enter guest name!")
                ok = False
            if am <= 0:
                st.error("❌ Enter valid amount!")
                ok = False

            if ok:
                bn = gen_bill_no()
                pn = sp.split(" — ")[0] if " — " in sp else sp

                bill_data = {
                    "bill_no": bn, "manual_bill_no": mbl, "bill_book_no": bb,
                    "devotee_type": "enrolled" if dt == "Enrolled Devotee" else "guest",
                    "devotee_id": did if dt == "Enrolled Devotee" else None,
                    "guest_name": d_name if dt == "Guest" else None,
                    "guest_address": d_addr if dt == "Guest" else None,
                    "guest_mobile": d_mob if dt == "Guest" else None,
                    "guest_whatsapp": d_wa if dt == "Guest" else None,
                    "pooja_type": pn, "amount": am, "bill_date": str(bd)
                }

                res = db_insert("bills", bill_data)

                if res:
                    st.success(f"✅ Bill Generated Successfully: **{bn}**")
                    st.balloons()

                    amman_img = get_amman_image()

                    # BEAUTIFUL BILL DISPLAY
                    st.markdown(f"""
                    <div class="bill-receipt">
                        <div class="bill-header">
                            <img src="{amman_img}" style="width:40px;height:40px;border-radius:50%;
                                position:absolute;left:10px;top:5px;border:2px solid #ff6b35;object-fit:cover;">
                            <h2 style="color:#8B0000;margin:0;">🛕 Sree Bhadreshwari Amman Temple</h2>
                            <p style="margin:3px 0;color:#5a1a00;">🙏 அம்மே நாராயணா ..தேவி நாராயணா 🙏</p>
                            <img src="{amman_img}" style="width:40px;height:40px;border-radius:50%;
                                position:absolute;right:10px;top:5px;border:2px solid #ff6b35;object-fit:cover;">
                        </div>
                        <table style="width:100%;margin:15px 0;border-collapse:collapse;">
                            <tr><td style="padding:6px;width:35%;"><b>📄 Bill No:</b></td><td>{bn}</td></tr>
                            <tr><td style="padding:6px;"><b>📝 Manual Bill:</b></td><td>{mbl or 'N/A'}</td></tr>
                            <tr><td style="padding:6px;"><b>📖 Book No:</b></td><td>{bb or 'N/A'}</td></tr>
                            <tr><td style="padding:6px;"><b>📅 Date:</b></td><td>{bd}</td></tr>
                            <tr><td colspan="2"><hr style="border:1px dashed #ddd"></td></tr>
                            <tr><td style="padding:6px;"><b>👤 Name:</b></td><td>{d_name}</td></tr>
                            <tr><td style="padding:6px;"><b>🏠 Address:</b></td><td>{d_addr}</td></tr>
                            <tr><td style="padding:6px;"><b>📱 Mobile:</b></td><td>{d_mob}</td></tr>
                            <tr><td style="padding:6px;"><b>📲 WhatsApp:</b></td><td>{d_wa}</td></tr>
                            <tr><td colspan="2"><hr style="border:1px dashed #ddd"></td></tr>
                            <tr><td style="padding:6px;"><b>🙏 Pooja:</b></td><td><b>{pn}</b></td></tr>
                            <tr><td style="padding:6px;"><b>💰 Amount:</b></td>
                                <td style="font-size:1.5em;color:#11998e;font-weight:bold;">₹ {am:,.2f}</td></tr>
                        </table>
                        <div style="text-align:center;border-top:3px solid #ff6b35;padding-top:10px;">
                            <p style="color:#666;margin:0;">🙏 Thank you! Amme Narayana .. Devi Narayana 🙏</p>
                        </div>
                    </div>""", unsafe_allow_html=True)

                    st.markdown("---")
                    st.markdown("### 📤 Download & Share")

                    dl1, dl2, dl3 = st.columns(3)

                    # PDF DOWNLOAD
                    with dl1:
                        if PDF_AVAILABLE:
                            try:
                                amman_pdf = get_amman_for_pdf()
                                pdf = generate_bill_pdf(bn, mbl, bb, bd, d_name, d_addr, d_mob, pn, am, amman_base64=amman_pdf)
                                st.download_button(
                                    "📥 Download PDF Bill",
                                    data=pdf,
                                    file_name=f"Bill_{bn}.pdf",
                                    mime="application/pdf",
                                    use_container_width=True
                                )
                            except Exception as ex:
                                st.warning(f"PDF error: {ex}")
                        else:
                            st.info("📄 PDF not available")

                    # WHATSAPP SEND
                    with dl2:
                        wa_number = d_wa or d_mob
                        if wa_number:
                            wa_msg = build_bill_whatsapp_message(bn, bd, d_name, pn, am, mbl, bb)
                            wa_link = make_whatsapp_link(wa_number, wa_msg)
                            st.markdown(f"""
                            <a href="{wa_link}" target="_blank" class="wa-btn" style="display:block;text-align:center;">
                                📲 Send Bill via WhatsApp
                            </a>
                            """, unsafe_allow_html=True)
                        else:
                            st.info("📱 No WhatsApp/Mobile number to send")

                    # COPY BILL TEXT
                    with dl3:
                        bill_text = f"Bill: {bn}\nDate: {bd}\nName: {d_name}\nPooja: {pn}\nAmount: ₹{am:,.2f}"
                        st.text_area("📋 Copy Bill Text", value=bill_text, height=100, key="copy_bill")

    # ---- BILL HISTORY TAB ----
    with tab2:
        st.markdown("### 📋 Bill History")

        # Filters
        hf1, hf2, hf3 = st.columns(3)
        with hf1:
            h_from = st.date_input("📅 From", value=date.today() - timedelta(30), key="hist_from")
        with hf2:
            h_to = st.date_input("📅 To", value=date.today(), key="hist_to")
        with hf3:
            h_search = st.text_input("🔍 Search (name/bill)", key="hist_search")

        all_bills = sorted(
            db_select("bills", gte_filters={"bill_date": h_from}, lte_filters={"bill_date": h_to}),
            key=lambda x: x.get('created_at', ''), reverse=True
        )

        for b in all_bills:
            bname = b.get('guest_name', '') or ''
            bwn = b.get('guest_whatsapp', '') or b.get('guest_mobile', '') or ''
            baddr = b.get('guest_address', '') or ''
            bmob = b.get('guest_mobile', '') or ''

            if b.get('devotee_type') == 'enrolled' and b.get('devotee_id'):
                dd = db_select("devotees", "name,mobile_no,whatsapp_no,address", filters={"id": b['devotee_id']})
                if dd:
                    bname = dd[0].get('name', '')
                    bmob = dd[0].get('mobile_no', '')
                    bwn = dd[0].get('whatsapp_no', '') or dd[0].get('mobile_no', '')
                    baddr = dd[0].get('address', '')

            # Apply search
            if h_search:
                if h_search.lower() not in f"{bname} {b.get('bill_no', '')} {b.get('pooja_type', '')}".lower():
                    continue

            with st.expander(f"🧾 {b.get('bill_no', '')} | {bname} | ₹{b.get('amount', 0):,.2f} | {b.get('bill_date', '')}"):
                # Bill details
                info1, info2 = st.columns(2)
                with info1:
                    st.write(f"**📄 Bill No:** {b.get('bill_no', '')}")
                    st.write(f"**📝 Manual:** {b.get('manual_bill_no', 'N/A')}")
                    st.write(f"**📖 Book:** {b.get('bill_book_no', 'N/A')}")
                    st.write(f"**📅 Date:** {b.get('bill_date', '')}")
                with info2:
                    st.write(f"**👤 Name:** {bname}")
                    st.write(f"**📱 Mobile:** {bmob}")
                    st.write(f"**🙏 Pooja:** {b.get('pooja_type', '')}")
                    st.write(f"**💰 Amount:** ₹{float(b.get('amount', 0)):,.2f}")

                # Action buttons
                act1, act2, act3 = st.columns(3)

                with act1:
                    if PDF_AVAILABLE:
                        try:
                            pdf_data = generate_bill_pdf(
                                b.get('bill_no', ''), b.get('manual_bill_no', ''),
                                b.get('bill_book_no', ''), b.get('bill_date', ''),
                                bname, baddr, bmob, b.get('pooja_type', ''),
                                b.get('amount', 0), amman_base64=get_amman_for_pdf()
                            )
                            st.download_button(
                                "📥 Download PDF",
                                data=pdf_data,
                                file_name=f"Bill_{b.get('bill_no', '')}.pdf",
                                mime="application/pdf",
                                key=f"pdf_{b['id']}"
                            )
                        except:
                            st.info("PDF N/A")

                with act2:
                    if bwn:
                        wa_msg = build_bill_whatsapp_message(
                            b.get('bill_no', ''), b.get('bill_date', ''),
                            bname, b.get('pooja_type', ''), b.get('amount', 0),
                            b.get('manual_bill_no', ''), b.get('bill_book_no', '')
                        )
                        wa_link = make_whatsapp_link(bwn, wa_msg)
                        st.markdown(
                            f'<a href="{wa_link}" target="_blank" class="wa-btn-small">'
                            f'📲 Send WhatsApp</a>',
                            unsafe_allow_html=True
                        )
                    else:
                        st.caption("No WhatsApp number")

                with act3:
                    if st.session_state.user_role == 'admin':
                        if st.button("🗑️ Delete", key=f"del_bill_{b['id']}"):
                            db_delete("bills", "id", b['id'])
                            st.rerun()


# ============================================================
# PAGE: EXPENSES
# ============================================================
def page_expenses():
    render_page_header("💸 Expenses", "Track temple expenses")
    t1, t2 = st.tabs(["➕ Add", "📋 History"])
    with t1:
        with st.form("ef", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                etn = [e['name'] for e in db_select("expense_types", "name")] or ["Misc"]
                et = st.selectbox("Type", etn)
                ea = st.number_input("Amount", min_value=0.0, step=10.0)
            with c2:
                ed = st.date_input("Date")
                edesc = st.text_area("Description", height=80)
            if st.form_submit_button("💾 Save", use_container_width=True):
                if ea > 0:
                    db_insert("expenses", {"expense_type": et, "amount": ea, "description": edesc, "expense_date": str(ed)})
                    st.rerun()
    with t2:
        exps = sorted(db_select("expenses"), key=lambda x: x.get('expense_date', ''), reverse=True)
        if exps:
            st.metric("Total", f"₹ {sum(float(e.get('amount', 0)) for e in exps):,.2f}")
            st.dataframe(pd.DataFrame([{
                "Date": e.get('expense_date', ''), "Type": e.get('expense_type', ''),
                "Amount": f"₹{float(e.get('amount', 0)):,.2f}", "Desc": e.get('description', '')
            } for e in exps]), use_container_width=True, hide_index=True)


# ============================================================
# PAGE: REPORTS
# ============================================================
def page_reports():
    render_page_header("📊 Reports", "Financial reports & analysis")
    rc1, rc2, rc3 = st.columns(3)
    with rc1: period = st.selectbox("Period", ["Daily", "Weekly", "Monthly", "Yearly", "Custom"])
    t = date.today()
    if period == "Custom":
        with rc2: sd = st.date_input("From", value=t - timedelta(30))
        with rc3: ed = st.date_input("To", value=t)
    else:
        sd, ed = get_period_dates(period)
    bills = db_select("bills", gte_filters={"bill_date": sd}, lte_filters={"bill_date": ed})
    exps = db_select("expenses", gte_filters={"expense_date": sd}, lte_filters={"expense_date": ed})
    ti = sum(float(b.get('amount', 0)) for b in bills)
    te = sum(float(e.get('amount', 0)) for e in exps)
    mc1, mc2, mc3 = st.columns(3)
    with mc1: st.markdown(f'<div class="metric-card income"><h3>Income</h3><h2>₹{ti:,.2f}</h2></div>', unsafe_allow_html=True)
    with mc2: st.markdown(f'<div class="metric-card expense"><h3>Expenses</h3><h2>₹{te:,.2f}</h2></div>', unsafe_allow_html=True)
    with mc3: st.markdown(f'<div class="metric-card balance"><h3>Balance</h3><h2>₹{ti - te:,.2f}</h2></div>', unsafe_allow_html=True)
    if bills:
        df = pd.DataFrame([{"Bill": b.get('bill_no', ''), "Date": b.get('bill_date', ''), "Pooja": b.get('pooja_type', ''), "Amount": float(b.get('amount', 0))} for b in bills])
        st.dataframe(df, use_container_width=True, hide_index=True)
        st.download_button("📥 CSV", df.to_csv(index=False), "income.csv")
    if bills or exps:
        st.bar_chart(pd.DataFrame({"Cat": ["Income", "Expenses"], "₹": [ti, te]}).set_index("Cat"))


# ============================================================
# PAGE: ASSETS WITH BARCODE
# ============================================================
def page_assets():
    render_page_header("🏷️ Assets", "Manage & Generate Barcodes")
    t1, t2, t3 = st.tabs(["➕ Add", "📋 List", "🏷️ Barcode Generator"])

    with t1:
        with st.form("af", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                at = st.text_input("🏷️ Tag *")
                an = st.text_input("📦 Name *")
                sn = st.text_input("🔢 Serial")
            with c2:
                dn = st.text_input("🙏 Donor")
                dd = st.date_input("📅 Date", min_value=MIN_DATE, max_value=MAX_DATE)
                ai = st.file_uploader("📷 Image", type=['jpg', 'jpeg', 'png'])
            adesc = st.text_area("📝 Notes", height=60)
            auto_bc = st.checkbox("🏷️ Auto-generate barcode", value=True)
            if st.form_submit_button("✅ Add Asset", use_container_width=True):
                if at.strip() and an.strip():
                    data = {"asset_tag": at.strip(), "asset_name": an.strip(), "serial_no": sn, "donor_name": dn,
                            "donation_date": str(dd), "image_url": file_to_base64(ai), "description": adesc}
                    if auto_bc:
                        bc_img, _ = generate_barcode_image(at.strip())
                        if bc_img:
                            data['barcode_url'] = bc_img
                    if db_insert("assets", data):
                        st.success(f"✅ Asset '{an}' added!")
                        st.rerun()

    with t2:
        a_search = st.text_input("🔍 Search Assets", key="as")
        assets = db_select("assets")
        if a_search:
            assets = [a for a in assets if a_search.lower() in f"{a.get('asset_tag', '')} {a.get('asset_name', '')}".lower()]
        st.markdown(f"**Total: {len(assets)}**")
        for a in assets:
            with st.expander(f"🏷️ {a.get('asset_tag', '')} | {a.get('asset_name', '')}"):
                ac1, ac2, ac3 = st.columns([2, 2, 1])
                with ac1:
                    for l, k in [("Tag", "asset_tag"), ("Name", "asset_name"), ("Serial", "serial_no"), ("Donor", "donor_name"), ("Date", "donation_date")]:
                        v = a.get(k, 'N/A')
                        if v and str(v) != 'None':
                            st.write(f"**{l}:** {v}")
                with ac2:
                    bc_img, bc_bytes = generate_barcode_image(a.get('asset_tag', ''))
                    if bc_img:
                        st.markdown(f'<div class="barcode-container"><img src="{bc_img}" style="max-width:250px"></div>', unsafe_allow_html=True)
                    if a.get('image_url') and a['image_url'].startswith('data:'):
                        st.markdown(f'<img src="{a["image_url"]}" width="130" style="border-radius:10px">', unsafe_allow_html=True)
                with ac3:
                    tag = a.get('asset_tag', '')
                    _, bc_dl = generate_barcode_image(tag)
                    if bc_dl:
                        st.download_button("📥 PNG", data=bc_dl, file_name=f"barcode_{tag}.png", mime="image/png", key=f"bp_{a['id']}")
                    if PDF_AVAILABLE:
                        bc_pdf = generate_asset_barcode_pdf(tag, a.get('asset_name', ''), bc_dl)
                        if bc_pdf:
                            st.download_button("📥 Label", data=bc_pdf, file_name=f"label_{tag}.pdf", mime="application/pdf", key=f"bl_{a['id']}")
                    if st.button("🗑️", key=f"da_{a['id']}"):
                        db_delete("assets", "id", a['id'])
                        st.rerun()

    with t3:
        st.markdown("### 🏷️ Bulk Barcode Generator")
        mode = st.radio("Mode", ["All Assets", "Custom Text"], horizontal=True)
        if mode == "All Assets":
            all_a = db_select("assets", "id,asset_tag,asset_name")
            if all_a and st.button("🏷️ Generate All", type="primary"):
                for i in range(0, len(all_a), 3):
                    cols = st.columns(3)
                    for j, asset in enumerate(all_a[i:i + 3]):
                        with cols[j]:
                            bc, _ = generate_barcode_image(asset['asset_tag'])
                            st.markdown(f'<div class="barcode-container"><p><b>{asset["asset_name"]}</b></p><img src="{bc}" style="max-width:200px"><p>{asset["asset_tag"]}</p></div>', unsafe_allow_html=True)
        else:
            ct = st.text_input("Enter text")
            if ct and st.button("🏷️ Generate", type="primary"):
                bc, bc_b = generate_barcode_image(ct)
                st.markdown(f'<div class="barcode-container"><img src="{bc}" style="max-width:350px"><p>{ct}</p></div>', unsafe_allow_html=True)
                if bc_b:
                    st.download_button("📥 Download", data=bc_b, file_name=f"barcode_{ct}.png", mime="image/png")


# ============================================================
# PAGE: SETTINGS
# ============================================================
def page_settings():
    render_page_header("⚙️ Settings", "Temple Configuration")
    t1, t2, t3, t4 = st.tabs(["🖼️ Amman Image", "🙏 Pooja Types", "💸 Expense Types", "📢 News Ticker"])

    with t1:
        st.markdown("### 🖼️ Amman Image Management")
        st.markdown("""
        Upload to display on: **Login page** (circle + background) • **Every page banner** • **Bill PDFs** • **Sidebar**
        """)
        current_img = get_amman_image()
        c1, c2 = st.columns([1, 2])
        with c1:
            st.markdown(f"""
            <div style="text-align:center;padding:20px;background:linear-gradient(135deg,#1a0a2e,#302b63);border-radius:15px;border:2px solid #f7c948;">
                <img src="{current_img}" style="width:150px;height:150px;border-radius:50%;border:4px solid #f7c948;object-fit:cover;box-shadow:0 0 30px rgba(247,201,72,0.4);">
                <p style="margin-top:10px;color:#f7c948;font-weight:bold;">Current Image</p>
            </div>""", unsafe_allow_html=True)
        with c2:
            st.info("📌 Recommended: Square image (500x500px+), JPG/PNG format")
            new_photo = st.file_uploader("Choose Amman image", type=['jpg', 'jpeg', 'png'], key="amman_settings")
            if new_photo:
                img_b64 = file_to_base64(new_photo)
                pc1, pc2, pc3 = st.columns(3)
                with pc1: st.markdown(f'<div style="text-align:center;"><p>Login</p><img src="{img_b64}" style="width:100px;height:100px;border-radius:50%;border:3px solid #ff6b35;object-fit:cover;"></div>', unsafe_allow_html=True)
                with pc2: st.markdown(f'<div style="text-align:center;"><p>Banner</p><img src="{img_b64}" style="width:50px;height:50px;border-radius:50%;border:2px solid #ff6b35;object-fit:cover;"></div>', unsafe_allow_html=True)
                with pc3: st.markdown(f'<div style="text-align:center;"><p>Sidebar</p><img src="{img_b64}" style="width:70px;height:70px;border-radius:50%;border:3px solid #ff6b35;object-fit:cover;"></div>', unsafe_allow_html=True)
                if st.button("✅ Save Amman Image", use_container_width=True, type="primary"):
                    save_amman_image_to_db(img_b64)
                    st.success("✅ Updated!")
                    st.balloons()
                    time.sleep(1)
                    st.rerun()
            if st.button("🔄 Reset to Default", use_container_width=True):
                st.session_state['custom_amman_photo'] = None
                try: db_delete("temple_settings", "key", "amman_image")
                except: pass
                st.success("✅ Reset!")
                st.rerun()

    with t2:
        for p in db_select("pooja_types"):
            c1, c2 = st.columns([5, 1])
            with c1: st.write(f"🙏 **{p['name']}** — ₹{p.get('amount', 0)}")
            with c2:
                if st.button("🗑️", key=f"dp_{p['id']}"): db_delete("pooja_types", "id", p['id']); st.rerun()
        with st.form("apt", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1: nn = st.text_input("Name")
            with c2: na = st.number_input("Amount", min_value=0.0, step=10.0)
            if st.form_submit_button("➕ Add"):
                if nn.strip(): db_insert("pooja_types", {"name": nn.strip(), "amount": na}); st.rerun()

    with t3:
        for e in db_select("expense_types"):
            c1, c2 = st.columns([5, 1])
            with c1: st.write(f"💸 **{e['name']}**")
            with c2:
                if st.button("🗑️", key=f"de_{e['id']}"): db_delete("expense_types", "id", e['id']); st.rerun()
        with st.form("aet", clear_on_submit=True):
            nn = st.text_input("Name")
            if st.form_submit_button("➕ Add"):
                if nn.strip(): db_insert("expense_types", {"name": nn.strip()}); st.rerun()

    with t4:
        for n in db_select("news_ticker"):
            c1, c2, c3 = st.columns([4, 1, 1])
            with c1: st.write(f"{'🟢' if n.get('is_active') else '🔴'} {n['message']}")
            with c2:
                if st.button("Toggle", key=f"tn_{n['id']}"): db_update("news_ticker", {"is_active": not n.get('is_active', True)}, "id", n['id']); st.rerun()
            with c3:
                if st.button("🗑️", key=f"dn_{n['id']}"): db_delete("news_ticker", "id", n['id']); st.rerun()
        with st.form("an", clear_on_submit=True):
            nm = st.text_input("Message")
            if st.form_submit_button("➕ Add"):
                if nm.strip(): db_insert("news_ticker", {"message": nm.strip(), "is_active": True}); st.rerun()


# ============================================================
# PAGE: USERS
# ============================================================
def page_users():
    render_page_header("👥 Users", "User Management")
    if st.session_state.user_role != 'admin':
        st.error("Admin only!")
        return
    t1, t2 = st.tabs(["➕ Create", "📋 List"])
    with t1:
        with st.form("cu", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1: nu = st.text_input("Username"); np_ = st.text_input("Password", type="password")
            with c2: cp = st.text_input("Confirm", type="password"); nr = st.selectbox("Role", ["user", "admin"])
            if st.form_submit_button("➕ Create", use_container_width=True):
                if nu and np_ and np_ == cp and not db_select("users", filters={"username": nu}):
                    db_insert("users", {"username": nu, "password_hash": np_, "role": nr})
                    st.success(f"✅ User '{nu}' created!")
                    st.rerun()
                elif np_ != cp:
                    st.error("Passwords don't match!")
    with t2:
        for u in db_select("users"):
            c1, c2 = st.columns([5, 1])
            with c1: st.write(f"{'👑' if u.get('role') == 'admin' else '👤'} **{u['username']}** ({u.get('role', '')})")
            with c2:
                if u['username'] != 'admin':
                    if st.button("🗑️", key=f"du_{u['id']}"): db_delete("users", "id", u['id']); st.rerun()


# ============================================================
# PAGE: SAMAYA VAKUPPU
# ============================================================
def page_samaya():
    render_page_header("📚 Samaya Vakuppu", "Student Management")
    t1, t2 = st.tabs(["➕ Add", "📋 List"])
    with t1:
        with st.form("sv", clear_on_submit=True):
            c1, c2 = st.columns(2)
            with c1:
                sn = st.text_input("Name *")
                sd = st.date_input("DOB", value=date(2010, 1, 1))
                spn = st.text_input("Parent Name")
                spt = st.selectbox("Parent Type", ["Father", "Mother"])
                sa = st.text_area("Address", height=60)
            with c2:
                sbn = st.text_input("Bond No")
                sbd = st.date_input("Bond Date")
                sbk = st.text_input("Bank")
                sbr = st.text_input("Branch")
                sbf = st.file_uploader("Bond Scan", type=['jpg', 'jpeg', 'png', 'pdf'], key="sv_bond")
                sph = st.file_uploader("Photo", type=['jpg', 'jpeg', 'png'], key="sv_photo")
            if st.form_submit_button("✅ Add", use_container_width=True):
                if sn.strip():
                    db_insert("samaya_vakuppu", {
                        "student_name": sn.strip(), "dob": str(sd), "address": sa,
                        "parent_name": spn, "parent_type": spt,
                        "bond_no": sbn, "bond_issue_date": str(sbd),
                        "bond_issuing_bank": sbk, "branch_of_bank": sbr,
                        "scanned_bond_url": file_to_base64(sbf),
                        "photo_url": file_to_base64(sph)
                    })
                    st.rerun()
    with t2:
        for s in db_select("samaya_vakuppu"):
            with st.expander(f"👤 {s['student_name']}"):
                for l, k in [("Name", "student_name"), ("DOB", "dob"), ("Parent", "parent_name"), ("Bond", "bond_no"), ("Bank", "bond_issuing_bank")]:
                    st.write(f"**{l}:** {s.get(k, 'N/A')}")
                if s.get('photo_url') and s['photo_url'].startswith('data:'):
                    st.markdown(f'<img src="{s["photo_url"]}" width="120" style="border-radius:10px">', unsafe_allow_html=True)
                if st.button("🗑️", key=f"ds_{s['id']}"):
                    db_delete("samaya_vakuppu", "id", s['id'])
                    st.rerun()


# ============================================================
# PAGE: THIRUMANA MANDAPAM (Enhanced with Scan Copy Upload)
# ============================================================
def page_thirumana():
    render_page_header("💒 Thirumana Mandapam", "Bond Management with Scan Copies")

    t1, t2 = st.tabs(["➕ Add New", "📋 Records"])

    with t1:
        st.markdown("### ➕ Add New Thirumana Mandapam Record")

        with st.form("tm_form", clear_on_submit=True):
            c1, c2 = st.columns(2)

            with c1:
                tn = st.text_input("👤 Name *", placeholder="Enter name")
                ta = st.text_area("🏠 Address", height=80, placeholder="Full address")
                tm_mob = st.text_input("📱 Mobile No", placeholder="Mobile number")
                tm_wa = st.text_input("📲 WhatsApp No", placeholder="WhatsApp number")

            with c2:
                tb = st.text_input("📄 Bond No", placeholder="Bond number")
                td = st.date_input("📅 Bond Issue Date", min_value=MIN_DATE, max_value=MAX_DATE)
                tam = st.number_input("💰 Amount (₹)", min_value=0.0, step=100.0)
                tnb = st.number_input("📊 No of Bonds", min_value=0, step=1)

            st.markdown("---")
            st.markdown("#### 📎 Upload Documents")

            doc1, doc2, doc3 = st.columns(3)
            with doc1:
                ts_bond = st.file_uploader(
                    "📜 Scan Copy of Bond *",
                    type=['jpg', 'jpeg', 'png', 'pdf'],
                    key="tm_bond_scan",
                    help="Upload scanned copy of the bond document"
                )
            with doc2:
                ts_receipt = st.file_uploader(
                    "🧾 Receipt Copy",
                    type=['jpg', 'jpeg', 'png', 'pdf'],
                    key="tm_receipt_scan",
                    help="Upload receipt if available"
                )
            with doc3:
                tp = st.file_uploader(
                    "📷 Photo",
                    type=['jpg', 'jpeg', 'png'],
                    key="tm_photo",
                    help="Upload photo of the person"
                )

            tm_notes = st.text_area("📝 Additional Notes", height=60, placeholder="Any additional information...")

            if st.form_submit_button("✅ Save Record", use_container_width=True):
                if tn.strip():
                    record_data = {
                        "name": tn.strip(),
                        "address": ta,
                        "bond_no": tb,
                        "bond_issued_date": str(td),
                        "amount": tam,
                        "no_of_bonds": tnb,
                        "scan_copy_url": file_to_base64(ts_bond),
                        "photo_url": file_to_base64(tp),
                    }
                    result = db_insert("thirumana_mandapam", record_data)
                    if result:
                        st.success(f"✅ Record for '{tn}' saved successfully!")
                        st.balloons()
                        st.rerun()
                else:
                    st.error("❌ Please enter name!")

    with t2:
        st.markdown("### 📋 Thirumana Mandapam Records")

        # Search
        tm_search = st.text_input("🔍 Search records", key="tm_search", placeholder="Search by name or bond no...")

        records = db_select("thirumana_mandapam")

        if tm_search:
            records = [r for r in records if tm_search.lower() in f"{r.get('name', '')} {r.get('bond_no', '')}".lower()]

        st.markdown(f"**Total Records: {len(records)}** | **Total Amount: ₹{sum(float(r.get('amount', 0)) for r in records):,.2f}**")

        for r in records:
            with st.expander(f"👤 {r.get('name', '')} | 📄 Bond: {r.get('bond_no', 'N/A')} | ₹{float(r.get('amount', 0)):,.2f}"):

                info1, info2 = st.columns(2)

                with info1:
                    st.markdown("#### 📋 Details")
                    for l, k in [
                        ("👤 Name", "name"),
                        ("🏠 Address", "address"),
                        ("📄 Bond No", "bond_no"),
                        ("📅 Issue Date", "bond_issued_date"),
                        ("💰 Amount", "amount"),
                        ("📊 No of Bonds", "no_of_bonds"),
                    ]:
                        val = r.get(k, 'N/A')
                        if k == 'amount' and val != 'N/A':
                            val = f"₹{float(val):,.2f}"
                        st.write(f"**{l}:** {val}")

                with info2:
                    st.markdown("#### 📎 Documents")

                    # Show scan copy of bond
                    if r.get('scan_copy_url') and r['scan_copy_url'].startswith('data:'):
                        st.markdown("**📜 Bond Scan Copy:**")
                        if 'pdf' in r['scan_copy_url']:
                            st.info("📄 PDF document uploaded. Click to view.")
                            # Extract and provide download
                            try:
                                pdf_data = base64.b64decode(r['scan_copy_url'].split(',')[1])
                                st.download_button(
                                    "📥 Download Bond PDF",
                                    data=pdf_data,
                                    file_name=f"bond_{r.get('bond_no', 'doc')}.pdf",
                                    mime="application/pdf",
                                    key=f"dl_bond_{r['id']}"
                                )
                            except:
                                pass
                        else:
                            st.markdown(f"""
                            <div class="scan-preview">
                                <img src="{r['scan_copy_url']}" alt="Bond Scan">
                            </div>""", unsafe_allow_html=True)
                    else:
                        st.caption("📜 No bond scan uploaded")

                    # Show photo
                    if r.get('photo_url') and r['photo_url'].startswith('data:'):
                        st.markdown("**📷 Photo:**")
                        st.markdown(f'<img src="{r["photo_url"]}" width="150" style="border-radius:10px;">', unsafe_allow_html=True)

                # Upload new scan copy for existing record
                with st.expander("📎 Upload/Update Scan Copy", expanded=False):
                    new_scan = st.file_uploader(
                        "Upload new bond scan copy",
                        type=['jpg', 'jpeg', 'png', 'pdf'],
                        key=f"new_scan_{r['id']}"
                    )
                    if new_scan:
                        scan_b64 = file_to_base64(new_scan)
                        if st.button("💾 Save Scan Copy", key=f"save_scan_{r['id']}"):
                            db_update("thirumana_mandapam", {"scan_copy_url": scan_b64}, "id", r['id'])
                            st.success("✅ Scan copy updated!")
                            st.rerun()

                    new_photo = st.file_uploader(
                        "Upload/Update photo",
                        type=['jpg', 'jpeg', 'png'],
                        key=f"new_photo_{r['id']}"
                    )
                    if new_photo:
                        photo_b64 = file_to_base64(new_photo)
                        if st.button("💾 Save Photo", key=f"save_photo_{r['id']}"):
                            db_update("thirumana_mandapam", {"photo_url": photo_b64}, "id", r['id'])
                            st.success("✅ Photo updated!")
                            st.rerun()

                # Edit & Delete
                edit_col, del_col = st.columns(2)
                with edit_col:
                    if st.button("✏️ Edit", key=f"edit_tm_{r['id']}"):
                        st.session_state[f"edit_tm_{r['id']}"] = not st.session_state.get(f"edit_tm_{r['id']}", False)
                        st.rerun()
                with del_col:
                    if st.button("🗑️ Delete", key=f"del_tm_{r['id']}"):
                        db_delete("thirumana_mandapam", "id", r['id'])
                        st.rerun()

                # Edit form
                if st.session_state.get(f"edit_tm_{r['id']}", False):
                    with st.form(f"edit_tm_form_{r['id']}"):
                        ec1, ec2 = st.columns(2)
                        with ec1:
                            e_name = st.text_input("Name", value=r.get('name', ''), key=f"etm_n_{r['id']}")
                            e_addr = st.text_area("Address", value=r.get('address', ''), key=f"etm_a_{r['id']}")
                            e_bond = st.text_input("Bond No", value=r.get('bond_no', ''), key=f"etm_b_{r['id']}")
                        with ec2:
                            e_amt = st.number_input("Amount", value=float(r.get('amount', 0)), key=f"etm_am_{r['id']}")
                            e_nb = st.number_input("No of Bonds", value=int(r.get('no_of_bonds', 0)), key=f"etm_nb_{r['id']}")
                        if st.form_submit_button("💾 Update"):
                            db_update("thirumana_mandapam", {
                                "name": e_name, "address": e_addr, "bond_no": e_bond,
                                "amount": e_amt, "no_of_bonds": e_nb
                            }, "id", r['id'])
                            st.session_state[f"edit_tm_{r['id']}"] = False
                            st.success("✅ Updated!")
                            st.rerun()


# ============================================================
# SIDEBAR
# ============================================================
def render_sidebar():
    with st.sidebar:
        amman_img = get_amman_image()
        st.markdown(f"""
        <div class="sidebar-amman"><img src="{amman_img}" alt="Amman"></div>
        <div style="text-align:center;padding:8px;background:linear-gradient(135deg,#ff6b35,#f7c948);border-radius:10px;margin-bottom:10px;">
            <p style="color:#5a1a00;margin:0;font-weight:700;font-size:0.75em;">Sree Bhadreshwari Amman<br>Temple Management</p>
        </div>
        <div style="color:#ccc;padding:5px 10px;font-size:0.8em;">
            👤 <b style="color:#f7c948">{st.session_state.username}</b> ({st.session_state.user_role})
        </div>""", unsafe_allow_html=True)
        st.markdown("---")

        pages = [
            ("🏠 Dashboard", "Dashboard"), ("👥 Devotees", "Devotees"),
            ("🧾 Billing", "Billing"), ("💸 Expenses", "Expenses"),
            ("📊 Reports", "Reports"), ("🏷️ Assets", "Assets"),
            ("📚 Samaya Vakuppu", "Samaya"), ("💒 Thirumana Mandapam", "Thirumana"),
            ("⚙️ Settings", "Settings"), ("👥 Users", "Users"),
        ]
        for l, p in pages:
            if p == "Users" and st.session_state.user_role != 'admin':
                continue
            if st.button(l, key=f"n_{p}", use_container_width=True):
                st.session_state.current_page = p
                st.rerun()

        st.markdown("---")
        if st.button("🚪 Logout", key="lo", use_container_width=True):
            for k in ['logged_in', 'username', 'user_role', 'current_page']:
                st.session_state[k] = defaults[k]
            st.rerun()

        st.markdown('<div style="text-align:center;padding:15px 0;color:#555;font-size:0.65em;">v3.1 🙏 அம்மே நாராயணா 🙏</div>', unsafe_allow_html=True)


# ============================================================
# MAIN
# ============================================================
def main():
    amman_img = get_amman_image()
    st.markdown(get_custom_css(amman_img), unsafe_allow_html=True)

    if not st.session_state.logged_in:
        page_login()
    else:
        render_sidebar()
        pm = {
            "Dashboard": page_dashboard, "Devotees": page_devotee_enrollment,
            "Billing": page_billing, "Expenses": page_expenses,
            "Reports": page_reports, "Assets": page_assets,
            "Samaya": page_samaya, "Thirumana": page_thirumana,
            "Settings": page_settings, "Users": page_users,
        }
        pm.get(st.session_state.current_page, page_dashboard)()


if __name__ == "__main__":
    main()
